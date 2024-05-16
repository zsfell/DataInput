from flask import Flask, request, render_template, redirect, url_for
import openpyxl

app = Flask(__name__)

excel_file = 'data.xlsx'
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["Data 1", "Data 2", "Data 3", "Data 4", "Data 5", "Data 6", "Data 7", "Data 8", "Data 9"])
    workbook.save(excel_file)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/submit', methods=['POST'])
def submit():
    data = [
        request.form['data1'].strip(),
        request.form['data2'].strip(),
        request.form['data3'].strip(),
        request.form['data4'].strip(),
        request.form['data5'].strip(),
        request.form['data6'].strip(),
        request.form['data7'].strip(),
        request.form['data8'].strip(),
        request.form['data9'].strip()
    ]
    
    data = [d if d.isdigit() else None for d in data]
    
    if all(d is None for d in data):
        return redirect(url_for('index'))

    last_row = sheet.max_row
    
    if last_row == 1 or all(sheet.cell(row=last_row, column=col).value is None for col in range(1, 10)):
        last_row += 1

    for col, value in enumerate(data, start=1):
        if value is not None:
            cell = sheet.cell(row=last_row, column=col)
            value = int(value)
            if cell.value is None:
                cell.value = value
            elif isinstance(cell.value, int):
                cell.value += value

    workbook.save(excel_file)
    return redirect(url_for('index'))

if __name__ == '__main__':
    app.run(debug=True)
